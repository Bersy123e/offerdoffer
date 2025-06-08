import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Optional
import logging
import uuid

from logger import setup_logger

logger = setup_logger()

class ProposalGenerator:
    def __init__(self, output_dir: str = "proposals"):
        """
        Инициализация генератора коммерческих предложений.
        
        Args:
            output_dir: Папка для сохранения готовых предложений
        """
        self.output_dir = output_dir
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """Создание папки нужно для сохранения файлов без ошибок."""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"Создана папка для предложений: {self.output_dir}")
    
    def generate(self, products: List[Dict], quantity: int = 10) -> str:
        """
        Генерация коммерческого предложения в Excel.
        
        Args:
            products: Список товаров для включения в предложение
            quantity: Количество по умолчанию (или остаток, если меньше)
            
        Returns:
            Путь к созданному Excel файлу
        """
        try:
            logger.info(f"Генерация предложения с {len(products)} товарами")
            
            # Создание нового файла Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Коммерческое предложение"
            
            # Добавление заголовков для структурирования документа
            self._add_headers(ws)
            
            # Добавление товаров из найденного списка
            self._add_products(ws, products, quantity)
            
            # Добавление итогов для расчета общей стоимости
            self._add_total(ws, len(products) + 3)  # +3 для строк заголовка
            
            # Форматирование для профессионального вида
            self._format_worksheet(ws, len(products) + 4)  # +4 для заголовка и итогов
            
            # Уникальное имя файла предотвращает перезапись
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            uid = str(uuid.uuid4())[:8]
            filename = f"КП_{timestamp}_{uid}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            # Сохранение готового файла
            wb.save(filepath)
            logger.info(f"Предложение сохранено: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка при генерации предложения: {str(e)}")
            # Перебрасываем исключение для обработки на верхнем уровне
            raise
    
    def _add_headers(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """
        Добавление заголовков нужно для структурирования документа.
        
        Args:
            ws: Лист Excel
        """
        # Заголовок документа для идентификации
        ws.merge_cells('A1:E1')
        ws['A1'] = "Коммерческое предложение"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Дата нужна для отслеживания актуальности
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y')}"
        ws['A2'].alignment = Alignment(horizontal='right')
        
        # Заголовки колонок для понимания структуры данных
        headers = ["№", "Наименование товара", "Цена (руб)", "Кол-во", "Сумма (руб)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def _add_products(self, ws: openpyxl.worksheet.worksheet.Worksheet, products: List[Dict], default_quantity: int):
        """
        Добавление товаров из поиска в таблицу предложения.
        
        Args:
            ws: Лист Excel
            products: Список найденных товаров
            default_quantity: Количество по умолчанию
        """
        for i, product in enumerate(products, 1):
            # Количество ограничено остатком для избежания проблем с поставкой
            quantity = min(default_quantity, product.get('stock', default_quantity))
            
            # Расчет суммы для каждой позиции
            price = product.get('price', 0)
            total = price * quantity
            
            # Заполнение строки данными товара
            ws.cell(row=i+3, column=1).value = i  # Номер для навигации
            ws.cell(row=i+3, column=2).value = product.get('name', '')  # Название товара
            ws.cell(row=i+3, column=3).value = price  # Цена за единицу
            ws.cell(row=i+3, column=4).value = quantity  # Количество к заказу
            ws.cell(row=i+3, column=5).value = total  # Общая стоимость позиции
            
            # Выравнивание для улучшения читаемости
            ws.cell(row=i+3, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=i+3, column=2).alignment = Alignment(vertical='center', wrap_text=True)
            ws.cell(row=i+3, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=i+3, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=i+3, column=5).alignment = Alignment(horizontal='right')
    
    def _add_total(self, ws: openpyxl.worksheet.worksheet.Worksheet, row: int):
        """
        Добавление итогов нужно для расчета общей стоимости заказа.
        
        Args:
            ws: Лист Excel
            row: Номер строки для итогов
        """
        # Объединение ячеек для надписи "Итого"
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "Итого:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        
        # Формула автоматически пересчитывает сумму при изменениях
        last_row = row - 1
        ws[f'E{row}'] = f'=SUM(E4:E{last_row})'
        ws[f'E{row}'].font = Font(bold=True)
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
    
    def _format_worksheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, last_row: int):
        """
        Форматирование нужно для профессионального внешнего вида документа.
        
        Args:
            ws: Лист Excel
            last_row: Номер последней строки
        """
        # Ширина колонок подбирается для оптимального отображения контента
        column_widths = {
            1: 5,   # Номер - узкая колонка
            2: 50,  # Наименование - широкая для длинных названий
            3: 15,  # Цена - средняя для чисел
            4: 10,  # Количество - узкая для чисел
            5: 15,  # Сумма - средняя для итогов
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Границы нужны для визуального разделения данных
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(3, last_row + 1):
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
        
        # Форматирование чисел для корректного отображения валюты
        for row in range(4, last_row + 1):
            # Колонка цены
            price_cell = ws.cell(row=row, column=3)
            price_cell.number_format = '#,##0.00'
            
            # Колонка суммы
            total_cell = ws.cell(row=row, column=5)
            total_cell.number_format = '#,##0.00' 